#! /usr/bin/env python3
# -*- Encoding: utf-8 -*-

import sqlite3
from tkinter import *
import tkinter.ttk as ttk
import datetime
from tkinter import messagebox
import os
import sys
import PyPDF2
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph
#from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill



# git
# utworzenie połączenia z bazą przechowywaną na dysku
# lub w pamięci (':memory:')
con = sqlite3.connect('baza_Amber.db')

# dostęp do kolumn przez indeksy i przez nazwy
con.row_factory = sqlite3.Row

# utworzenie obiektu kursora
cur = con.cursor()





# lista zagnieżdzona przekazywania stanow kontrolek
l=200     #parametry listy-ilosc zagniezdzen
x=200    #parametry listy- ilosc zmiennych(str) w zagniezdzeniu
State=[]
for s in range(l):
    sub_list=[""]
    for y in range(x):

        sub_list.append(int(y))

        State.append(sub_list)

n=-1  #zmienna odpow. za poruszanie po bazie



# Klasa okna glownego
class Application(Frame):


    def __init__(self,master,State,n):
        """Inicjuję ramkę"""

        dt = datetime.datetime.now()
        self.now_d = dt.strftime("%Y-%m-%d")
        self.now_h = dt.strftime("%H:%M")
        self.lbl_alarm_nw_info=Label()
        self.lbl_dzial_nag_info=Label()
        self.lbl_wpi_term_info=Label()


        super(Application,self).__init__(master)
        self.grid()
        self.nr_fab(State)
        self.kod_prod(State)
        self.nr_zlec(State)
        self.identy_nr_fab(State)
        self.zagiecia_estetyka(State)
        self.filtry_uszczelki_silikonowanie(State)
        self.szczelnosc_membran_dlawic(State)
        self.dokrecenie_wymiennika(State)
        self.prowadzenie_przewodow_ssr(State)
        self.montaz_rodzielnicy_zabez_nadpradowe(State)
        self.montaz_nagrzewnicy(State)
        self.jakosc_polaczen_elektrycznych(State)
        self.wpiecie_przewodow_termikow(State)
        self.wentylatory_montaz_dzialanie(State)
        self.czujniki_temper_montaz_dzial(State)
        self.dzialanie_nagrzewnicy_elektrycznej (State)
        self.wyzwolenie_alarmu_nag_ele(State)
        self.montaz_uziemienia(State)
        self.konfiguracja_peryferii_jezyk(State)
        self.bloki_czas_pozw_grza(State)
        self.piktogramy_oznaczenia_jezyk(State)
        self.dokrecenie_srub_ogranicznikow(State)
        self.kompletnosc_dostawy(State)
        self.uwagi(State)
        self.okno_uwagi(State)
        self.podpis_kontrolera(State)
        self.btn_akcept()
        self.btn_drukuj()
        self.protocol_perm()
        self.n = n
        self.btn_date()
        self.btn_zrzut()






########################################################################################################################################################################################

    # poziom lini nr fabryczny


    def nr_fab(self,State):
        """numer fabryczny"""

        # utworz etykiete z zapytniem o nr fabryczny
        self.lbl_dist_0=Label(self)
        self.lbl_dist_0.grid(row = 0, column = 0 , padx=6)                       #dystans col 0

        self.lbl_nr_fab = Label(self, text ="Podaj nr fabryczny")
        self.lbl_nr_fab.grid(row = 0, column = 1 , sticky = W )

        var1 = StringVar()  # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1] == 0:
            var1.set("")
        else:
            var1.set(State[0][1])

        self.ent_nr_fab = Entry(self, textvariable=var1)
        self.ent_nr_fab.grid(row=0, column=4, ipadx=10)

        self.lbl_dist_2=Label(self)
        self.lbl_dist_2.grid(row = 0, column = 2 , padx=5)                               #dystans  col  2

        self.lbl_dist_3=Label(self)
        self.lbl_dist_3.grid(row = 0, column = 3 , padx=1)                               #dystans   col 3



        # utworz widzet Entry do przyjecia nr_fab

        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
        else:
            var1.set(State[0][1])


        self.ent_nr_fab= Entry(self,textvariable=var1)
        self.ent_nr_fab.grid(row=0, column = 4, ipadx=10 )

        self.lbl_dist_5=Label(self)
        self.lbl_dist_5.grid(row = 0, column = 5 , padx=4)                               #dystans col  5



        #utworz przyciski 'archiwum'5

        self.btn_up = Button(self,text= "UP", command=self.arch_UP )
        self.btn_up.grid(row = 0, column=6 , ipadx=20 , sticky=E,padx=(15,0) )

        self.btn_dn = Button(self,text= "DN", command=self.arch_DOWN )
        self.btn_dn.grid(row = 0, column=7 ,ipadx=20, sticky=W)




        self.lbl_dist_8=Label(self)
        self.lbl_dist_8.grid(row = 0, column = 5 , padx=4, pady=3)                               #dystans col  8


########################################################################################################################################################################################
    def onReturn(self,event):                                          # modół sterujący wyświetlaniem widzetów

        prod_cod=str(self.typ_ahu.get())
        if prod_cod[-3] == "1":
            messagebox.showinfo("info","Nagrzewnica elektryczna wbudowana")


        else:

            messagebox.showinfo("info","Bez nagrzewnicy elektrycznej wbudowanej")

            # ALARM
            self.lbl_alarm_nw_info=Label(self, text=(State[0][17]))
            self.lbl_alarm_nw_info.configure(text='Brak')
            self.lbl_alarm_nw_info.grid(row=16, column=7)
            State[0][17] = 'Brak'
            self.wyzwolenie_alarmu_nag_ele(State)

            #Działanie NW
            self.lbl_dzial_nag_info = Label(self, text=(State[0][16]))
            self.lbl_dzial_nag_info.configure(text='Brak')
            self.lbl_dzial_nag_info.grid(row=15, column=7)
            State[0][16] = 'Brak'
            self.dzialanie_nagrzewnicy_elektrycznej (State)

            #Wpięcie termików

            self.lbl_wpi_term_info = Label(self, text=(State[0][13]))
            self.lbl_wpi_term_info.configure(text='Brak')
            self.lbl_wpi_term_info.grid(row=12, column=7)
            State[0][13] = 'Brak'
            self.wpiecie_przewodow_termikow(State)

            #Montaż nagrzewnicy

            State[0][11]="Brak"
            self.montaz_nagrzewnicy(State)





    def kod_prod(self,State):

        self.nagrzewnica = TRUE

        self.czynnosc = Label(self, text ="Podaj kod produktu")
        self.czynnosc.grid(row = 1, column = 1 , sticky = W )



        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
        else:
            var1.set(State[0][2])


        self.typ_ahu= Entry(self,textvariable=var1)

        self.typ_ahu.bind("<Return>",self.onReturn)
        self.typ_ahu.grid(row=1, column = 4, ipadx=10 )



        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
            Time_label_txt=""
        else:
            var1.set(State[0][2])
            Time_label_txt ="Ostatnie badanie:"





         # Wyswietlanie komunikatu
        self.lbl_t1 = Label(self, text=Time_label_txt)
        self.lbl_t1.grid(row = 1, column =6, columnspan=2)










    #    self.czynnosc = Label(self, text ="Podaj kod produktu")
    #   self.czynnosc.grid(row = 1, column = 1 , sticky = W )



    #    var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

    #    if State[0][1]==0 :
    #        var1.set("")
    #        Time_label_txt=""
    #    else:
    #        var1.set(State[0][2])
    #        Time_label_txt ="Ostatnie badanie:"

    #    self.ent_kod_prod= Entry(self,textvariable=var1)
    #    self.ent_kod_prod.grid(row=1, column = 4)
     #  self.lbl_dist_9=Label(self)
     #   self.lbl_dist_9.grid(row = 1, column = 5,pady=3 )

        # Wyswietlanie komunikatu
     #   self.lbl_t1 = Label(self, text=Time_label_txt)
     #   self.lbl_t1.grid(row = 1, column =6, columnspan=2)







########################################################################################################################################################################################


    def nr_zlec(self,State):

        # utworz etykiete z zapytniem o zlecenie
        self.czynnosc = Label(self, text ="Podaj nr zlecenia")
        self.czynnosc.grid(row = 2, column = 1 , columnspan = 1, sticky = W )






        # utworz widzet Entry do przyjecia zlecenia


        var2=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera
        Time_txt=StringVar()
        t1=(str(State[0][27])+' '+str(State[0][28]))

        if State[0][1]==0 :
            var2.set("")
            t1=str(self.now_d+'  '+str(self.now_h))
            Time_txt.set(t1)

        else:
            var2.set(State[0][3])
            t1=(str(State[0][27])+' '+str(State[0][28]))
            Time_txt.set(t1)
            self.now_d = str(State[0][27])
            self.now_h = str(State[0][28])


        self.ent_nr_zlec= Entry(self,textvariable=var2)
        self.ent_nr_zlec.grid(row=2, column = 4, ipadx=10)

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 2, column = 5 , pady=3)




        self.lbl_t2 = Entry(self, text = Time_txt, state = DISABLED)
        self.lbl_t2.grid(row = 2, column =6, columnspan=2 )

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 28, column = 5 , pady=3)
        var3=StringVar()
        a=var3.set("wprowadz datę")
        self.ent_data= Entry(self, textvariable=var3 )
        self.ent_data.grid(row=29, column = 1, ipadx=15)

        #contens3 = str(self.ent_nr_zlec.get())




    def btn_date(self):
        self.nowa_data = Button(self, text ="Nowa data",command = self.give_new_date)
        self.nowa_data.grid(row = 29, column = 3)

    def give_new_date(self):
        self.now_d = self.ent_data.get()






########################################################################################################################################################################################


    def identy_nr_fab(self,State):


        self.czynnosc = Label(self, text ="Identyfikacja , nalepka z nr. fabrycznym ")
        self.czynnosc.grid(row = 3, column = 1,sticky = W )
        self.ident_tab_znam = StringVar()
        self.ident_tab_znam.set(State[0][4])



        Radiobutton(self,
                    text =  "Tak",
                    variable = self.ident_tab_znam,
                    value = "Pozytyw",
                    ).grid(row = 3, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.ident_tab_znam,
                    value = "Negatyw",
                    ).grid(row = 3, column = 4, sticky=E)



    # Wyswietlanie stanu   z bazy
        if State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][4] )
            self.lbl_czynnosc_info.grid(row = 3, column = 7)

#######################################################################################################################################################################################



    def zagiecia_estetyka(self, State):

        self.czynnosc = Label(self, text ="Zagięcia, estetyka, rysy")
        self.czynnosc.grid(row = 4, column = 1,sticky = W )

        self.zag_estet = StringVar()

        self.zag_estet.set(State[0][5])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.zag_estet,
                    value = "Pozytyw",
                    ).grid(row = 4, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.zag_estet,
                    value = "Negatyw",
                    #command = self.update_text
                    ).grid(row = 4, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][5])
            self.lbl_czynnosc_info.grid(row = 4, column = 7)




#######################################################################################################################################################################################


    def filtry_uszczelki_silikonowanie(self, State):

        self.czynnosc = Label(self, text ="Filtry , uszczelki, silikonowanie ")
        self.czynnosc.grid(row = 5, column = 1,sticky = W )

        self.fil_uszcz_silik = StringVar()

        self.fil_uszcz_silik.set(State[0][6])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.fil_uszcz_silik,
                    value = "Pozytyw",
                    # command = self.update_text
                    ).grid(row = 5, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.fil_uszcz_silik,
                    value = "Negatyw",
                    #command = self.update_text
                    ).grid(row = 5, column = 4,sticky=E)


       # Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][6])
            self.lbl_czynnosc_info.grid(row = 5, column = 7)






######################################################################################################################################################################################




    def szczelnosc_membran_dlawic(self, State):

        self.lbl_czynnosc = Label(self, text ="Szczelność membran , dławic")
        self.lbl_czynnosc.grid(row = 6, column = 1,sticky = W )

        self.szczel_memb_dlawic= StringVar()

        self.szczel_memb_dlawic.set(State[0][7])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.szczel_memb_dlawic,
                    value = "Pozytyw",
                    ).grid(row = 6, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.szczel_memb_dlawic,
                    value = "Negatyw",
                    ).grid(row = 6, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][7])
            self.lbl_czynnosc_info.grid(row = 6, column = 7)


#######################################################################################################################################################################################


    def dokrecenie_wymiennika(self, State):


        self.lbl_czynnosc = Label(self, text ="Dokręcenie wymiennika, szczelność ")
        self.lbl_czynnosc.grid(row = 7, column = 1,sticky = W )

        self.dokr_wymienn = StringVar()

        self.dokr_wymienn.set(State[0][8])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.dokr_wymienn,
                    value = "Pozytyw",
                    ).grid(row = 7, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.dokr_wymienn,
                    value = "Negatyw",
                    ).grid(row = 7, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][8])
            self.lbl_czynnosc_info.grid(row = 7, column = 7)



########################################################################################################################################################################################


    def prowadzenie_przewodow_ssr(self, State):

        self.lbl_czynnosc = Label(self, text ="Prowadzenie przewodów, montaż SSR" )
        self.lbl_czynnosc.grid(row = 8, column = 1,sticky = W )

        self.prowadz_przewodow = StringVar()

        self.prowadz_przewodow.set(State[0][9])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.prowadz_przewodow,
                    value = "Pozytyw",
                    ).grid(row = 8, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.prowadz_przewodow,
                    value = "Negatyw",
                    ).grid(row = 8, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][9])
            self.lbl_czynnosc_info.grid(row = 8, column = 7)









########################################################################################################################################################################################


    def montaz_rodzielnicy_zabez_nadpradowe(self, State):


        self.lbl_czynnosc = Label(self, text ="Montaż rozdzielnicy , zabezp. nadprądowe" )
        self.lbl_czynnosc.grid(row = 9, column = 1,sticky = W )

        self.mont_rozdz_zabez_nadpr = StringVar()

        self.mont_rozdz_zabez_nadpr.set(State[0][10])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.mont_rozdz_zabez_nadpr,
                    value = "Pozytyw",
                    ).grid(row = 9, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.mont_rozdz_zabez_nadpr,
                    value = "Negatyw",
                    ).grid(row = 9, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][10])
            self.lbl_czynnosc_info.grid(row = 9, column = 7)


########################################################################################################################################################################################


    def montaz_nagrzewnicy (self, State):



        self.lbl_czynnosc = Label(self, text ="Montaż nagrzewnicy - monter" )
        self.lbl_czynnosc.grid(row = 10, column = 1,sticky = W )

        self.mon_nagrzewnicy = StringVar()

        self.combobox_mon_nag  = ttk.Combobox(self, textvariable=self.mon_nagrzewnicy)
        self.combobox_mon_nag.grid(row=10, column=4, columnspan=2, sticky=W)
        self.combobox_mon_nag['values'] = ('', 'Brak', 'Marian', 'Dariusz')
        self.combobox_mon_nag.current(0)
        a = State[0][11]
        if a == 'Brak' :
            self.combobox_mon_nag.current(1)
        if a == 'Marian':
            self.combobox_mon_nag.current(2)

        if a == 'Dariusz' :
            self.combobox_mon_nag.current(3)












########################################################################################################################################################################################



    def jakosc_polaczen_elektrycznych(self, State):


        self.lbl_czynnosc = Label(self, text ="Jakość połączeń elektrycznych")
        self.lbl_czynnosc.grid(row = 11, column = 1,sticky = W )

        self.jak_pol_ele = StringVar()

        self.jak_pol_ele.set(State[0][12])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.jak_pol_ele,
                    value = "Pozytyw",
                    ).grid(row = 11, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.jak_pol_ele,
                    value = "Negatyw",
                    ).grid(row = 11, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][12])
            self.lbl_czynnosc_info.grid(row = 11, column = 7)



########################################################################################################################################################################################


    def wpiecie_przewodow_termikow(self, State):


        self.lbl_czynnosc = Label(self, text ="Wpięcie przewodów teremików" )
        self.lbl_czynnosc.grid(row = 12, column = 1,sticky = W )

        self.wp_przew_term = StringVar()

        self.wp_przew_term.set(State[0][13])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.wp_przew_term,
                    value = "Pozytyw",
                    ).grid(row = 12, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.wp_przew_term,
                    value = "Negatyw",
                    ).grid(row = 12, column = 4,sticky=E)

        Radiobutton(self,
                    text="Brak",
                    variable=self.wp_przew_term,
                    value="Brak",
                    ).grid(row=12, column=6, sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_wpi_term_info = Label(self,text=State[0][13])
            self.lbl_wpi_term_info.grid(row = 12, column = 7)


####################################################################################################################################################################################


    def wentylatory_montaz_dzialanie(self,State):


        self.lbl_czynnosc = Label(self, text ="Wentylatory- montaż , działnie" )
        self.lbl_czynnosc.grid(row = 13, column = 1,sticky = W )

        self.went = StringVar()

        self.went.set(State[0][14])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.went,
                    value = "Pozytyw",
                    ).grid(row = 13, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.went,
                    value = "Negatyw",
                    ).grid(row = 13, column = 4,sticky=E)



# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][14])
            self.lbl_czynnosc_info.grid(row = 13, column = 7)


#######################################################################################################################################################################################

    def czujniki_temper_montaz_dzial(self,State):


        self.lbl_czynnosc = Label(self, text ="Czujniki temeratur -montaż, działanie" )
        self.lbl_czynnosc.grid(row = 14, column = 1,sticky = W )

        self.czujn_temp = StringVar()

        self.czujn_temp.set(State[0][15])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.czujn_temp,
                    value = "Pozytyw",
                    ).grid(row = 14, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.czujn_temp,
                    value = "Negatyw",
                    ).grid(row = 14, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][15])
            self.lbl_czynnosc_info.grid(row = 14, column = 7)


#######################################################################################################################################################################################



    def dzialanie_nagrzewnicy_elektrycznej (self, State):

        self.lbl_czynnosc = Label(self, text ="Działanie nagrzewnicy elektrycznej")
        self.lbl_czynnosc.grid(row = 15, column = 1,sticky = W )

        self.dzial_nw = StringVar()

        self.dzial_nw.set(State[0][16])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.dzial_nw,
                    value = "Pozytyw",
                    ).grid(row = 15, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.dzial_nw,
                    value = "Negatyw",
                    ).grid(row = 15, column = 4,sticky=E)

        Radiobutton(self,
                    text="Brak",
                    variable = self.dzial_nw,
                    value="Brak",
                    ).grid(row=15, column=6, sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_dzial_nag_info = Label(self,text=State[0][16])
            self.lbl_dzial_nag_info.grid(row = 15, column = 7)






#######################################################################################################################################################################################


    def wyzwolenie_alarmu_nag_ele (self, State):




            self.lbl_czynnosc = Label(self, text ="Wywołanie alarmu nagrzewnicy elektr." )
            self.lbl_czynnosc.grid(row = 16, column = 1,sticky = W )

            self.wyw_alarm_ne = StringVar()

            self.wyw_alarm_ne.set(State[0][17])


            Radiobutton(self,
                        text =  "Tak",
                        variable = self.wyw_alarm_ne,
                        value = "Pozytyw",
                        ).grid(row = 16, column = 4,sticky=W)

            Radiobutton(self,
                        text =  "Nie",
                        variable = self.wyw_alarm_ne,
                        value = "Negatyw",
                        ).grid(row = 16, column = 4,sticky=E)


            Radiobutton(self,
                         text =  "Brak",
                         variable = self.wyw_alarm_ne,
                         value = "Brak",
                         ).grid(row = 16, column = 6,sticky=E)




    # Wyswietlanie stanu  z bazy

            if  State[0][1]!=0 :
                self.lbl_alarm_nw_info = Label(self,text=State[0][17])
                self.lbl_alarm_nw_info.grid(row = 16, column = 7)


#######################################################################################################################################################################################




    def montaz_uziemienia (self, State):

        self.lbl_czynnosc = Label(self, text ="Montaż uziemienia" )
        self.lbl_czynnosc.grid(row = 17, column = 1,sticky = W )

        self.mont_uziem = StringVar()

        self.mont_uziem.set(State[0][18])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.mont_uziem,
                    value = "Pozytyw",
                    ).grid(row = 17, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.mont_uziem,
                    value = "Negatyw",
                    ).grid(row = 17, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][18])
            self.lbl_czynnosc_info.grid(row = 17, column = 7)



######################################################################################################################################################################################

    def konfiguracja_peryferii_jezyk(self, State):

        self.lbl_czynnosc = Label(self, text="Konfiguracja peryferii, język ")
        self.lbl_czynnosc.grid(row=18, column=1, sticky=W)

        self.konfig_per_jez = StringVar()

        self.konfig_per_jez.set(State[0][19])

        Radiobutton(self,
                    text="Tak",
                    variable=self.konfig_per_jez,
                    value="Pozytyw",
                    ).grid(row=18, column=4, sticky=W)

        Radiobutton(self,
                    text="Nie",
                    variable=self.konfig_per_jez,
                    value="Negatyw",
                    ).grid(row=18, column=4, sticky=E)

        # Wyswietlanie stanu  z bazy

        if State[0][1] != 0:
            self.lbl_czynnosc_info = Label(self, text=State[0][19])
            self.lbl_czynnosc_info.grid(row=18, column=7)


#######################################################################################################################################################################################

    def bloki_czas_pozw_grza(self, State):

        self.lbl_czynnosc = Label(self, text="Bloki czasowe, pozwolenie na grzanie")
        self.lbl_czynnosc.grid(row=19, column=1, sticky=W)

        self.blok_czas = StringVar()

        self.blok_czas.set(State[0][20])

        Radiobutton(self,
                    text="Tak",
                    variable=self.blok_czas,
                    value="Pozytyw",
                    ).grid(row=19, column=4, sticky=W)

        Radiobutton(self,
                    text="Nie",
                    variable=self.blok_czas,
                    value="Negatyw",
                    ).grid(row=19, column=4, sticky=E)

        # Wyswietlanie stanu  z bazy

        if State[0][1] != 0:
            self.lbl_czynnosc_info = Label(self, text=State[0][20])
            self.lbl_czynnosc_info.grid(row=19, column=7)




###############################################################################################################################################################################


    def piktogramy_oznaczenia_jezyk  (self, State):


        self.lbl_czynnosc = Label(self, text ="Piktogramy, oznaczenia, język" )
        self.lbl_czynnosc.grid(row = 20, column = 1,sticky = W )

        self.oznacz_jezyk = StringVar()

        self.oznacz_jezyk.set(State[0][21])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.oznacz_jezyk,
                    value = "Pozytyw",
                    ).grid(row = 20, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.oznacz_jezyk,
                    value = "Negatyw",
                    ).grid(row = 20, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][21])
            self.lbl_czynnosc_info.grid(row = 20, column = 7)






#######################################################################################################################################################################################


    def dokrecenie_srub_ogranicznikow (self, State):

        self.lbl_czynnosc = Label(self, text="Dokręcenie śrub ograniczników")
        self.lbl_czynnosc.grid(row=21, column=1, sticky=W)


        self.dokr_srub_ogr = StringVar()



        self.combobox_mon_nag = ttk.Combobox(self, textvariable=self.dokr_srub_ogr)
        self.combobox_mon_nag.grid(row=21, column=4, columnspan=2, sticky=W)
        self.combobox_mon_nag['values'] = ('', 'nie', 'Dariusz', 'Leszek')
        self.combobox_mon_nag.current(0)
        a = State[0][22]
        if a == 'nie':
            self.combobox_mon_nag.current(1)
        if a == 'Dariusz':
            self.combobox_mon_nag.current(2)

        if a == 'Leszek':
            self.combobox_mon_nag.current(3)



    #######################################################################################################################################################################################


    def kompletnosc_dostawy (self, State):

        self.lbl_czynnosc = Label(self, text="Kompletność dostawy")
        self.lbl_czynnosc.grid(row=22, column=1, sticky=W)

        self.komplet_dostawy = StringVar()

        self.combobox_mon_nag = ttk.Combobox(self, textvariable=self.komplet_dostawy)
        self.combobox_mon_nag.grid(row=22, column=4, columnspan=2, sticky=W)
        self.combobox_mon_nag['values'] = ('', 'nie', 'Dariusz', 'Leszek')
        self.combobox_mon_nag.current(0)

        a = State[0][23]
        if a == 'nie':
            self.combobox_mon_nag.current(1)
        if a == 'Dariusz':
            self.combobox_mon_nag.current(2)

        if a == 'Leszek':
            self.combobox_mon_nag.current(3)



    #######################################################################################################################################################################################

    def uwagi(self,State):

        self.lbl_czynnosc = Label(self,text ="Uwagi")
        self.lbl_czynnosc.grid(row = 23, column = 1 , sticky =W)  # dystans


        self.uwagi_bool = StringVar()
        self.uwagi_bool.set(State[0][24])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.uwagi_bool,
                    value = "Pozytyw",
                    ).grid(row = 23, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.uwagi_bool,
                    value = "Negatyw",
                    ).grid(row = 23, column = 4,sticky=E)

# Wyswietlanie stanu  z bazy


        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][24])
            self.lbl_czynnosc_info.grid(row = 23, column = 7)


#######################################################################################################################################################################################


    def okno_uwagi(self,State):

        var3=StringVar() # zmienna pomocnicza - ukrywanie wywswl - zera
        var3=State [0][25]
        self.uwagi_txt = Text(self, width = 65, height = 1, wrap=CHAR )
        self.uwagi_txt.grid(row = 24, column = 1, columnspan = 10 , sticky =W)

        if State[0][1]==0 :
            self.uwagi_txt.insert(END," ")
        else:
            self.uwagi_txt.insert(END,var3)

####################################################################################################################################################################################

#dystans

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 25, column = 1)


####################################################################################################################################################################################

    def podpis_kontrolera(self,State):

        self.podpis = StringVar()
        self.combobox_podpis = ttk.Combobox(self, textvariable = self.podpis)
        self.combobox_podpis.grid(row = 26, column = 3, columnspan = 2 , sticky =W)
        self.combobox_podpis['values'] = ('','Jarosław Olszewski', 'Piotr Tylak','Dominik Tanski')
        self.combobox_podpis.current(0)
        a=State[0][26]
        if a=='Jaroslaw Olszewski' or a=='Jarosław Olszewski':
            self.combobox_podpis.current(1)
        if a=='Piotr Tylak':
            self.combobox_podpis.current(2)
        if a=='Dominik Tański'or a=='Dominik Tanski':
            self.combobox_podpis.current(3)





######################################################################################################################################################################################


    #poziom lini przycisk akceptuj

    def btn_akcept(self):
        #utworz przycisk - akceptuj- poziom
        self.submit_bttn = Button(self, text ="Akceptuj", command = self.ostrzezenie_zapis)
        self.submit_bttn.grid(row = 26, column = 1)


    def ostrzezenie_zapis(self):
        if messagebox.askyesno("Zapis danych", "Czy zapisać dane ?"):
            self.akcept()

    def btn_drukuj(self):
        self.print_bttn = Button(self, text ="Drukuj",command = self.info_druk)
        self.print_bttn.grid(row = 26, column = 7)



    def info_druk(self):

        if messagebox.askyesno("Wydruk do PDF", "Czy zapisać i wydrukować do PDF ?"):
            self.control_list(State)


            self.filling_factory(State)
            self.akcept()


    # funkcja przycisku akeptuj
    def akcept (self):


        contens1 = str(self.ent_nr_fab.get())
        contens2 = str(self.typ_ahu.get())
        contens3 = str(self.ent_nr_zlec.get())
        id1 = self.ident_tab_znam.get()
        id2 = self.zag_estet.get()
        id3 = self.fil_uszcz_silik.get()
        id4 = self.szczel_memb_dlawic.get()
        id5 = self.dokr_wymienn.get()
        id6 = self.prowadz_przewodow.get()
        id7 = self.mont_rozdz_zabez_nadpr.get()
        id8 = self.mon_nagrzewnicy.get()
        id9 = self.jak_pol_ele.get()
        id10 = self.wp_przew_term.get()
        id11 = self.went.get()
        id12 = self.czujn_temp.get()
        id13 = self.dzial_nw.get()
        id14 = self.wyw_alarm_ne.get()
        id15 = self.mont_uziem.get()
        id16 = self.konfig_per_jez.get()
        id17 = self.blok_czas.get()
        id18 = self.oznacz_jezyk.get()
        id19 = self.dokr_srub_ogr.get()
        id20 = self.komplet_dostawy.get()
        id21 = self.uwagi_bool.get()
        id22 = self.uwagi_txt.get(1.0, END)
        id23 = self.podpis.get()
        id24 = str(self.now_d)
        id25 = str(self.now_h)
        cur.execute('INSERT INTO tab VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,DATE(?),?);',(contens1,contens2,contens3,\
        id1,id2,id3,id4,id5,id6,id7,id8,id9,id10,id11,id12,id13,id14,id15,id16,id17,id18,id19,id20,id21,id22,id23,id24,id25))

        messagebox.showinfo("Zapis danych", "Zapisano")
        con.commit()



    def btn_zrzut(self):
        self.print_bttn = Button(self, text ="Zrzut",command = self.zrzut)
        self.print_bttn.grid(row = 29, column = 7)


    def zrzut(self):

        fill_color='BFBFBF'#'a5a391'

        excelfile = openpyxl.load_workbook('C:\\Backup\\KCX_Lista_temp.xlsx')  # open a excel file with .xlsx format
        excelfile.get_sheet_names()  # get names of all spreadsheet in the file
        sheet1 = excelfile.get_sheet_by_name("2019")  # get the first spreadsheet by name
        nr_line = sheet1.max_row  # get the number of rows in the sheet



        uwag= sheet1.cell(row = (nr_line)+1, column = 5)

        uwag.value = self.uwagi_txt.get(1.0, END)





        uwag.alignment= Alignment(horizontal='center')

        a=self.uwagi_bool.get()
        messagebox.showinfo('sdsd',a)
        if a=='Pozytyw':
            wypelnienie=True
        else:
            wypelnienie=False

        if wypelnienie==True:
            uwag.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        numfab  = sheet1.cell(row = (nr_line)+1, column = 2 )
        numfab.value = str(self.ent_nr_fab.get())
        numfab.alignment= Alignment(horizontal='center')

        if wypelnienie==True:
            numfab.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")



        empty_01 = sheet1.cell(row = (nr_line)+1, column = 3 )
        if wypelnienie==True:
            empty_01.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        kod_prod  = sheet1.cell(row = (nr_line)+1, column = 4 )
        kod_prod.value = self.typ_ahu.get()
        kod_prod.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            kod_prod.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")




        Year,month,day=self.now_d.split('-')
        data_format= day + '-' + month + '-'+Year
        data= sheet1.cell(row = (nr_line)+1, column = 6)
        data.value=data_format
        data.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            data.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        #messagebox.showinfo('fddf',data_raw)
        zlec= sheet1.cell(row = (nr_line)+1, column = 7)
        zlec.value = int(self.ent_nr_zlec.get())
        zlec.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            zlec.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")






        try:
            #('C:\\Users\\jolszewski\Desktop\\Expenses01.xlsx')

            excelfile.save('C:\\Backup\\KCX_Lista_temp.xlsx')###
            messagebox.showinfo('info','Zrzut do Excel poprawny')
        except:

            messagebox.showinfo('info','Cos poszło nie tak')



    #  funkcja zwiększająca zmeinną sterującą bazą
    def arch_UP(self):
        self.n+=1
        self.arch(self.n)

    #  funkcja zmniejszająca zmeinną sterującą bazą
    def arch_DOWN(self):
        self.n-=1
        self.arch(self.n)


#########################################################################################################################################################################################
#  wydruk PDF

    def control_list(self,State):

      #  deklaracja zmiennych


        self.pt_contens1 = str(self.ent_nr_fab.get())
        self.pt_contens2 = self.typ_ahu.get()
        self.pt_contens3 = str(self.ent_nr_zlec.get())
        self.pt_id1 = self.ident_tab_znam.get()
        self.pt_id2 = self.zag_estet.get()
        self.pt_id3 = self.fil_uszcz_silik.get()
        self.pt_id4 = self.szczel_memb_dlawic.get()
        self.pt_id5 = self.dokr_wymienn.get()
        self.pt_id6 = self.prowadz_przewodow.get()
        self.pt_id7 = self.mont_rozdz_zabez_nadpr.get()
        self.pt_id8 = self.mon_nagrzewnicy.get()
        self.pt_id9 = self.jak_pol_ele.get()
        self.pt_id10 = self.wp_przew_term.get()
        self.pt_id11 = self.went.get()
        self.pt_id12 = self.czujn_temp.get()
        self.pt_id13 = self.dzial_nw.get()
        self.pt_id14 = self.wyw_alarm_ne.get()
        self.pt_id15 = self.mont_uziem.get()
        self.pt_id16 = self.konfig_per_jez.get()
        self.pt_id17 = self.blok_czas.get()
        self.pt_id18 = self.oznacz_jezyk.get()
        self.pt_id19 = self.dokr_srub_ogr.get()
        self.pt_id20 = self.komplet_dostawy.get()
        self.pt_id21 = self.uwagi_bool.get()
        self.pt_id22 = self.uwagi_txt.get("1.0", END)
        self.pt_id23 = self.podpis.get()
        self.pt_time = str(self.now_d+ " "+self.now_h)



        pdfmetrics.registerFont(TTFont('DejaMono', 'DejaVuSansMono.ttf'))

        doc = SimpleDocTemplate("__file__"+"../../listy_kontrolne/"+self.pt_contens1+"_lista_kontr_KCX.pdf", pagesize=A4 )

        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']

        elements = []







        #ocena sumaryczna
        self.summary_mark ="Negatyw"
        self.protocol_permission=False

        if self.pt_id1=="Pozytyw" and self.pt_id2=="Pozytyw" and self.pt_id3=="Pozytyw" and self.pt_id4=="Pozytyw"\
        and self.pt_id5=="Pozytyw" and self.pt_id6=="Pozytyw" and self.pt_id7=="Pozytyw" and self.pt_id7=="Pozytyw" \
        and self.pt_id9=="Pozytyw" and self.pt_id10!="Negatyw" and self.pt_id11=="Pozytyw" and self.pt_id12=="Pozytyw" and self.pt_id13!="Negatyw" \
        and self.pt_id14!="Negatyw" and self.pt_id15=="Pozytyw" and self.pt_id16=="Pozytyw" and self.pt_id17=="Pozytyw" and self.pt_id18=="Pozytyw" \
        and self.pt_id19!="nie" and self.pt_id20!="nie" and self.pt_id21=="Pozytyw" :
            self.summary_mark ="Pozytyw"
            self.protocol_permission=TRUE







        data= [['Logo','','Montaż aparatów AMBER            ZP3 SKOWARCZ' ],
               ['','','Lista kontrolna czynnosci sprawdzajacych AMBER'],
               ['Lp', 'Operacja', 'Lp', 'Wytyczne', 'Ocena'],
               ['1', 'Kontrola wizualna', '1.1', 'Identyfikacja, naklejka z nr. fab ',self.pt_id1],
               ['' , '', '1.2', 'Zagięcia, estetyka , rysy', self.pt_id2],
               ['', '' , '1.3', 'Filtry, uszczelki, silikonowanie', self.pt_id3],
               ['', '', '1.4', 'Szczelność membran, dławic', self.pt_id4],
               ['', '', '1.5', 'Dokręcenie wymiennika, szczelność', self.pt_id5],
               ['', '', '1.6', 'Prow. przewodów ,montaż SSR', self.pt_id6],
               ['', '', '1.7', 'Montaż rozdzielni, zabezp. nadprądowe', self.pt_id7],
               ['', '', '1.8', 'Montaż nagrzewnicy', self.pt_id8],
               ['', '', '1.9', 'Jakość połączeń elektrycznych', self.pt_id9],
               ['', '', '1.10', 'Wpięcie przewodów termików', self.pt_id10],
               ['2', 'Próby ruchowe','2.2', 'Wentylatory-montaż-działanie', self.pt_id11],
               ['', '', '2.3', 'Czujniki temperatur-montaż-działanie', self.pt_id12],
               ['', '', '2.4', 'Działanie nagrzewnicy elektrycznej', self.pt_id13],
               ['', '', '2.5', 'Wywołanie alarmu nagrzewnicy elektr.', self.pt_id14],
               ['', '', '2.6', 'Montaż uziemienia', self.pt_id15],
               ['', '', '2.7', 'Konfiguracja peryferii, język', self.pt_id16],
               ['', '', '2.8', 'Bloki czasowe', self.pt_id17],
               ['3', 'Kompletacja','3.1' , 'Piktogramy, oznacznia, język' , self.pt_id18],
               ['', '', '3.2', 'Dokręcenie ogranicników' , self.pt_id19],
               ['', '', '3.3', 'Pakowanie dostawy' , self.pt_id20],
               ['4', 'Uwagi',self.pt_id22,'', self.pt_id21],
               ['5', 'Nr zlecenia', str(self.pt_contens3), '' , ''],
               ['',  'Nr seryjny', str(self.pt_contens1), '' , ''],
               ['',  'Typ urzadzenia',str(self.pt_contens2), '' , ''],
               ['',  'Data badania KJ', self.pt_time, '' , ''],
               ['',  'Kontroler KJ', self.pt_id23, '' , ''],
               ['6',  'Ocena końcowa',self.summary_mark, '' , ''],


               ]



        t=Table(data,4*[1.3*inch], 30*[0.3*inch])
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),

                               #('VALIGN',(0,0),(0,-1),'TOP'),
                               #('TEXTCOLOR',(0,0),(0,-1),colors.blue),
                               ('ALIGN',(0,0),(-1,-1),'CENTER'),
                               ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                               #('TEXTCOLOR',(0,-1),(-1,-1),colors.green),
                               ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                               ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                               ('SPAN',(3,0),(4,0)),
                               ('SPAN',(2,1),(4,1)),
                               ('SPAN',(2,0),(4,0)),
                               ('SPAN',(0,0),(1,1)),
                               ('SPAN',(0,3),(0,12)),#
                               ('SPAN',(1,3),(1,12)),#
                               ('SPAN',(0,13),(0,19)),#
                               ('SPAN',(1,13),(1,19)),#
                               ('SPAN',(0,20),(0,22)),#
                               ('SPAN',(1,20),(1,22)),#
                               ('SPAN',(2,23),(3,23)),
                               ('SPAN',(2,24),(4,24)),
                               ('SPAN',(0,24),(0,28)),
                               ('SPAN',(2,25),(4,25)),
                               ('SPAN',(2,26),(4,26)),
                               ('SPAN',(2,27),(4,27)),
                               ('SPAN',(2,28),(4,28)),
                               ('SPAN',(2,29),(4,29)),
                               ('SPAN',(2,30),(4,30)),
                               ('FONTNAME',(0,0),(-1,-1),'DejaMono'),
                               ('FONTSIZE',(2,1),(2,4),12),
                               ('FONTSIZE',(2,0),(2,4),10),


                               ]))

        t._argW[0]=0.5*inch
        t._argW[1] = 1.5*inch
        t._argW[2] = 0.5*inch
        t._argW[3] = 3.2*inch
        t._argW[4] = 1.*inch
        t._argH[1]=0.4*inch
        t._argH[23]=0.5*inch
        t._argH[29]=0.5*inch
        elements.append(t)
        # write the document to disk
        doc.build(elements)
###############################################################################################################################################################################

    def filling_factory(self,State):

        pdfmetrics.registerFont(TTFont('DejaMono', 'DejaVuSansMono.ttf'))

        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']

        doc = SimpleDocTemplate("__file__"+"../../protokol/filling_protokol_KCX.pdf", pagesize=A4 )




        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']




        elements = []

        kom_NW=0






        if self.wyw_alarm_ne.get()== "Pozytyw":
            kom_NW="Sprawdzono w działaniu"

        if self.wyw_alarm_ne.get()== "Brak":
            kom_NW="BRAK"





        czas=str(self.now_d+ " "+self.now_h)

        data= [['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','',self.typ_ahu.get()+" \ "+ str(self.ent_nr_zlec.get()) ,'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','',str(self.ent_nr_fab.get()),'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','        '+str(kom_NW),'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''], ###
               ['','','','','','','','','',''],
               ['','','','','','','','','',''], ###
               ['','','','','','','','','',''],
               ['','','',self.podpis.get(),'','','','',czas,''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ]



        t=Table(data,10*[0.7*inch], 60*[0.1445*inch])
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),

             ('VALIGN',(0,49),(9,49),'TOP'),
             ('ALIGN',(0,28),(9,28),'CENTER'),
             ('FONTNAME',(0,0),(-1,-1),'DejaMono')



                               ]))


        elements.append(t)
        # write the document to disk
        doc.build(elements)
        self.protocol()







##################################################################################################################################################################################

    def protocol_perm(self):

        self.protocol_permission=True

##################################################################################################################################################################################

    def protocol(self):

        if self.protocol_permission==TRUE :

                row_protocol_KCX = open('__file__"+"../../protokol/raw_protocol_KCX.pdf', 'rb')
                pdfReader = PyPDF2.PdfFileReader(row_protocol_KCX)
                row_protocol_KCX_p0 = pdfReader.getPage(0)

                filling_protocol_KCX = (open("__file__"+"../../protokol/filling_protokol_KCX.pdf", 'rb'))
                pdfFillingReader = PyPDF2.PdfFileReader(filling_protocol_KCX)
                filling_protocol_KCX_p0 = pdfFillingReader.getPage(0)

                row_protocol_KCX_p0.mergePage(filling_protocol_KCX_p0)
                pdfWriter=PyPDF2.PdfFileWriter()
                pdfWriter.addPage(row_protocol_KCX_p0)

                resultPdfFile = open("__file__"+"../../protocols/"+self.pt_contens1+"_protocol_KCX.pdf", 'wb')
                pdfWriter.write(resultPdfFile)
                row_protocol_KCX .close()
                resultPdfFile.close()

                messagebox.showinfo('Wydruk protokołu', 'Protokół zostanie wydrukowany')
        else:
                messagebox.showinfo('Wydruk protokołu', 'Protokół nie zostanie wydrukowany')



##################################################################################################################################################################################

    def warning (self):
        messagebox.showwarning("Warning", "Popraw wpisana liczę")
        python = sys.executable
        os.execl(python, python, * sys.argv)
###################################################################################################################################################################################

    # funkcja przycisku archiwizacja - THE END
    def arch (self,n):
        State=[]


        cur.execute(
            """
            SELECT ID,nr_fabr,kod_prod, nr_zlec,ident_tab_znam,zag_estet,fil_uszcz_silik,szczel_memb_dlawic,dokr_wymienn,prowadz_przewodow,\
            mont_rozdz_zabez_nadpr,mon_nagrzewnicy,jak_pol_ele,wp_przew_term,went,czujn_temp,dzial_nw,wyw_alarm_ne,mont_uziem,konfig_per_jez,blok_czas,\
            oznacz_jezyk,dokr_srub_ogr,komplet_dostawy,uwagi_bool,uwagi_txt,podpis,now_d,now_h FROM tab

            """)
        State_Train = cur.fetchall()
        State=State_Train[self.n:]


        self.grid_remove()


        app1 = Application(root,State,self.n)
        return State,app1,self.n

      #  except:

        #    messagebox.showinfo("Zapis danych", "Poza zakresem")

         #   python = sys.executable
         #   os.execl(python, python, * sys.argv)


# czesc glowna


root = Tk()
root.title("Amber 1")
root.geometry("660x770")



app = Application(root,State,n)



root.mainloop()

